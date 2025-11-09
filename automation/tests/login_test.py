# from pages.login_page import LoginPage
# from utils.excel_logger import ExcelLogger
# from utils.browser_setup import start_browser
# from selenium.webdriver.support import expected_conditions as EC

# def run_login_tests():
#     driver, wait = start_browser()
#     log = ExcelLogger()
#     login = LoginPage(driver, wait)

#     # ✅ Valid Login
#     try:
#         login.login("admin", "admin@1234")
#         wait.until(EC.url_contains("dashboard"))
#         log.log("LOGIN", "TC_LOGIN_01", "Valid Login", "PASS")
#     except Exception as e:
#         log.log("LOGIN", "TC_LOGIN_01", "Valid Login", "FAIL", str(e))

#     driver.quit()

# from automation.utils.browser_setup import get_driver

# def run_login_tests():
#     driver = get_driver()
#     driver.get("https://example.com/login")
#     print("Login test executed")
#     driver.quit()

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import traceback
import datetime
import os
from openpyxl import Workbook, load_workbook
import time


class TestOrionLogin:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        self.wait = WebDriverWait(self.driver, 10)
        self.url = "http://192.168.2.221:3000"
        self.excel_path = os.path.join(os.path.dirname(__file__), "test_results.xlsx")

        # Initialize Excel file
        if not os.path.exists(self.excel_path):
            self.create_excel()

    def create_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Test Case ID", "Description", "Status", "Reason", "Timestamp"])
        wb.save(self.excel_path)

    def log_result(self, tc_id, description, status, reason=""):
        """Append test results to Excel file."""
        wb = load_workbook(self.excel_path)
        ws = wb.active
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([tc_id, description, status, reason.strip(), timestamp])
        wb.save(self.excel_path)

        # Print to terminal
        print(f"{tc_id}: {description}")
        print(f" Status: {status}")
        if reason:
            print(f" Reason: {reason.strip()}")

    def open_login_page(self):
        self.driver.get(self.url)

    def find(self, by, value):
        return self.wait.until(EC.presence_of_element_located((by, value)))

    def login(self, username, password, remember=False, toggle=False):
        """Reusable login helper."""
        self.open_login_page()

        username_input = self.find(By.ID, "username")
        password_input = self.driver.find_element(By.ID, "password")
        login_button = self.driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[1]/div/form/div[3]/button')

        # clear and fill inputs
        username_input.clear()
        password_input.clear()
        if username:
            username_input.send_keys(username)
        if password:
            password_input.send_keys(password)

        # Remember me (if applicable)
        if remember:
            try:
                remember_checkbox = self.driver.find_element(By.XPATH, '//input[@type="checkbox"]')
                if not remember_checkbox.is_selected():
                    remember_checkbox.click()
            except NoSuchElementException:
                pass

        # Toggle password visibility
        if toggle:
            try:
                toggle_icon = self.driver.find_element(By.XPATH, '//button[contains(@aria-label,"toggle password")]')
                toggle_icon.click()
                time.sleep(1)
                toggle_icon.click()
            except NoSuchElementException:
                pass

        login_button.click()

    # ---------------- TEST CASES ---------------- #

    def test_TC_LOGIN_01(self):
        """Login with correct username/email and password"""
        tc_id, desc = "TC_LOGIN_01", "Login with correct username/email and password"
        try:
            self.login("admin", "admin@1234")
            self.wait.until(EC.url_contains("orion-dashboard"))
            print(f" {tc_id}: PASS")
            self.log_result(tc_id, desc, "PASS")
        except Exception as e:
            self.log_result(tc_id, desc, "FAIL", str(e))

    def test_TC_LOGIN_02(self):
        """Login with incorrect username/email and valid password"""
        tc_id, desc = "TC_LOGIN_02", "Login with incorrect username/email and valid password"
        try:
            self.login("wronguser", "admin@1234")
            time.sleep(2)
            if "login" in self.driver.current_url:
                raise Exception("Invalid credentials accepted.")
            self.log_result(tc_id, desc, "PASS")
        except Exception as e:
            self.log_result(tc_id, desc, "FAIL", str(e))

    def test_TC_LOGIN_03(self):
        """Login with blank username/email"""
        tc_id, desc = "TC_LOGIN_03", "Login with blank username/email"
        try:
            self.login("", "admin@1234")
            time.sleep(2)
            if "login" not in self.driver.current_url:
                raise Exception("Login succeeded with blank username.")
            self.log_result(tc_id, desc, "PASS")
        except Exception as e:
            self.log_result(tc_id, desc, "FAIL", str(e))

    def test_TC_LOGIN_04(self):
        """Login with blank password"""
        tc_id, desc = "TC_LOGIN_04", "Login with blank password"
        try:
            self.login("admin", "")
            time.sleep(2)
            if "login" not in self.driver.current_url:
                raise Exception("Login succeeded with blank password.")
            self.log_result(tc_id, desc, "PASS")
        except Exception as e:
            self.log_result(tc_id, desc, "FAIL", str(e))

    def test_TC_LOGIN_05(self):
        """Login with both fields blank"""
        tc_id, desc = "TC_LOGIN_05", "Login with both fields blank and press Sign in"
        try:
            self.login("", "")
            time.sleep(2)
            if "login" not in self.driver.current_url:
                raise Exception("Login succeeded with both fields blank.")
            self.log_result(tc_id, desc, "PASS")
        except Exception as e:
            self.log_result(tc_id, desc, "FAIL", str(e))

    def test_TC_LOGIN_06(self):
        """Check remember me functionality"""
        tc_id, desc = "TC_LOGIN_06", "Check 'Remember me', login successfully, logout, and revisit"
        try:
            self.login("admin", "admin@1234", remember=True)
            self.wait.until(EC.url_contains("orion-dashboard"))
            # simulate logout
            self.driver.get(f"{self.url}/logout")
            self.open_login_page()
            self.log_result(tc_id, desc, "PASS")
        except Exception as e:
            self.log_result(tc_id, desc, "FAIL", str(e))

    def test_TC_LOGIN_07(self):
        """Toggle password visibility"""
        tc_id, desc = "TC_LOGIN_07", "Toggle password visibility using eye icon"
        try:
            self.login("admin", "admin@1234", toggle=True)
            self.log_result(tc_id, desc, "PASS")
        except Exception as e:
            self.log_result(tc_id, desc, "FAIL", str(e))
            
            
    def test_TC_LOGIN_08(self):
        """Login with correct username 'admin' and password 'admin@123'"""
        tc_id, desc = "TC_LOGIN_08", "Login with correct username 'admin' and password 'admin@123'"
        try:
            self.login("admin", "admin@123")
            # Wait until dashboard loads
            self.wait.until(EC.url_contains("orion-dashboard"))
            self.log_result(tc_id, desc, "PASS")
        except Exception as e:
            self.log_result(tc_id, desc, "FAIL", str(e))



if __name__ == "__main__":
    test_suite = TestOrionLogin()
    # Run all test cases
    test_suite.test_TC_LOGIN_01()
    test_suite.test_TC_LOGIN_02()
    test_suite.test_TC_LOGIN_03()
    test_suite.test_TC_LOGIN_04()
    test_suite.test_TC_LOGIN_05()
    test_suite.test_TC_LOGIN_06()
    test_suite.test_TC_LOGIN_07()
    test_suite.test_TC_LOGIN_08()
